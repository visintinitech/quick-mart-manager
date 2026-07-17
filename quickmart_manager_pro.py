"""
QuickMart Manager PRO - Sistema de Gestión para Supermercado
Versión 2.0 - Con GUI, autenticación, PDF, filtros, exportación, descuentos, compras
Desarrollado por: VisintiniTech
"""

📦 Dependencias externas (instalar con pip)

pip install reportlab openpyxl

import sqlite3
import hashlib
import csv
import os
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment

# ============================================
# CONFIGURACIÓN DE BASE DE DATOS
# ============================================

DB_NAME = "quickmart_pro.db"

def conectar_bd():
    return sqlite3.connect(DB_NAME)

def crear_tablas():
    conn = conectar_bd()
    c = conn.cursor()

    # Tabla de usuarios (autenticación)
    c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            rol TEXT NOT NULL CHECK (rol IN ('admin', 'vendedor'))
        )
    ''')

    # Tabla de proveedores
    c.execute('''
        CREATE TABLE IF NOT EXISTS proveedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            telefono TEXT,
            email TEXT,
            direccion TEXT
        )
    ''')

    # Tabla de productos
    c.execute('''
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            precio REAL NOT NULL CHECK (precio >= 0),
            stock INTEGER NOT NULL CHECK (stock >= 0),
            categoria TEXT,
            proveedor_id INTEGER,
            FOREIGN KEY (proveedor_id) REFERENCES proveedores(id) ON DELETE SET NULL
        )
    ''')

    # Tabla de clientes
    c.execute('''
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            telefono TEXT,
            email TEXT,
            direccion TEXT
        )
    ''')

    # Tabla de ventas (cabecera)
    c.execute('''
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            cliente_id INTEGER,
            total REAL NOT NULL CHECK (total >= 0),
            descuento REAL DEFAULT 0,
            FOREIGN KEY (cliente_id) REFERENCES clientes(id) ON DELETE SET NULL
        )
    ''')

    # Tabla de detalle de ventas
    c.execute('''
        CREATE TABLE IF NOT EXISTS detalle_venta (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venta_id INTEGER NOT NULL,
            producto_id INTEGER NOT NULL,
            cantidad INTEGER NOT NULL CHECK (cantidad > 0),
            precio_unitario REAL NOT NULL CHECK (precio_unitario >= 0),
            subtotal REAL NOT NULL CHECK (subtotal >= 0),
            FOREIGN KEY (venta_id) REFERENCES ventas(id) ON DELETE CASCADE,
            FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE RESTRICT
        )
    ''')

    # Tabla de compras a proveedores (reposición)
    c.execute('''
        CREATE TABLE IF NOT EXISTS compras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            proveedor_id INTEGER,
            total REAL NOT NULL CHECK (total >= 0),
            FOREIGN KEY (proveedor_id) REFERENCES proveedores(id) ON DELETE SET NULL
        )
    ''')

    # Tabla de detalle de compras
    c.execute('''
        CREATE TABLE IF NOT EXISTS detalle_compra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compra_id INTEGER NOT NULL,
            producto_id INTEGER NOT NULL,
            cantidad INTEGER NOT NULL CHECK (cantidad > 0),
            costo_unitario REAL NOT NULL CHECK (costo_unitario >= 0),
            subtotal REAL NOT NULL CHECK (subtotal >= 0),
            FOREIGN KEY (compra_id) REFERENCES compras(id) ON DELETE CASCADE,
            FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE RESTRICT
        )
    ''')

    # Tabla de promociones/descuentos
    c.execute('''
        CREATE TABLE IF NOT EXISTS promociones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK (tipo IN ('porcentaje', 'fijo')),
            valor REAL NOT NULL CHECK (valor >= 0),
            fecha_inicio TEXT,
            fecha_fin TEXT,
            producto_id INTEGER,
            FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE
        )
    ''')

    # Insertar usuario admin por defecto si no existe
    c.execute("SELECT * FROM usuarios WHERE username = 'admin'")
    if not c.fetchone():
        password_hash = hashlib.sha256("admin123".encode()).hexdigest()
        c.execute("INSERT INTO usuarios (username, password, rol) VALUES (?, ?, ?)",
                  ('admin', password_hash, 'admin'))

    conn.commit()
    conn.close()

# ============================================
# FUNCIONES DE AUTENTICACIÓN
# ============================================

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def autenticar_usuario(username, password):
    conn = conectar_bd()
    c = conn.cursor()
    c.execute("SELECT id, rol, password FROM usuarios WHERE username = ?", (username,))
    usuario = c.fetchone()
    conn.close()
    if usuario and usuario[2] == hash_password(password):
        return usuario[0], usuario[1]  # id, rol
    return None, None

def cambiar_contrasena(usuario_id, nueva_password):
    conn = conectar_bd()
    c = conn.cursor()
    c.execute("UPDATE usuarios SET password = ? WHERE id = ?", (hash_password(nueva_password), usuario_id))
    conn.commit()
    conn.close()

# ============================================
# FUNCIONES DE NEGOCIO (CRUD y operaciones)
# ============================================

# --- Proveedores ---
def agregar_proveedor(nombre, telefono, email, direccion):
    conn = conectar_bd()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO proveedores (nombre, telefono, email, direccion) VALUES (?, ?, ?, ?)",
                  (nombre, telefono, email, direccion))
        conn.commit()
        return True
    except sqlite3.Error as e:
        print(e)
        return False
    finally:
        conn.close()

def obtener_proveedores():
    conn = conectar_bd()
    c = conn.cursor()
    c.execute("SELECT id, nombre, telefono, email, direccion FROM proveedores ORDER BY nombre")
    datos = c.fetchall()
    conn.close()
    return datos

# --- Clientes ---
def agregar_cliente(nombre, telefono, email, direccion):
    conn = conectar_bd()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO clientes (nombre, telefono, email, direccion) VALUES (?, ?, ?, ?)",
                  (nombre, telefono, email, direccion))
        conn.commit()
        return True
    except sqlite3.Error:
        return False
    finally:
        conn.close()

def obtener_clientes():
    conn = conectar_bd()
    c = conn.cursor()
    c.execute("SELECT id, nombre, telefono, email, direccion FROM clientes ORDER BY nombre")
    datos = c.fetchall()
    conn.close()
    return datos

# --- Productos ---
def agregar_producto(codigo, nombre, descripcion, precio, stock, categoria, proveedor_id):
    conn = conectar_bd()
    c = conn.cursor()
    try:
        c.execute("""INSERT INTO productos (codigo, nombre, descripcion, precio, stock, categoria, proveedor_id)
                     VALUES (?, ?, ?, ?, ?, ?, ?)""",
                  (codigo, nombre, descripcion, precio, stock, categoria, proveedor_id))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    except sqlite3.Error:
        return False
    finally:
        conn.close()

def obtener_productos(filtros=None):
    """Obtiene productos con filtros opcionales: nombre, categoria, proveedor_id, precio_min, precio_max, stock_min"""
    conn = conectar_bd()
    c = conn.cursor()
    query = """
        SELECT p.id, p.codigo, p.nombre, p.descripcion, p.precio, p.stock, p.categoria, pr.nombre
        FROM productos p
        LEFT JOIN proveedores pr ON p.proveedor_id = pr.id
        WHERE 1=1
    """
    params = []
    if filtros:
        if filtros.get('nombre'):
            query += " AND p.nombre LIKE ?"
            params.append(f"%{filtros['nombre']}%")
        if filtros.get('categoria'):
            query += " AND p.categoria LIKE ?"
            params.append(f"%{filtros['categoria']}%")
        if filtros.get('proveedor_id'):
            query += " AND p.proveedor_id = ?"
            params.append(filtros['proveedor_id'])
        if filtros.get('precio_min') is not None:
            query += " AND p.precio >= ?"
            params.append(filtros['precio_min'])
        if filtros.get('precio_max') is not None:
            query += " AND p.precio <= ?"
            params.append(filtros['precio_max'])
        if filtros.get('stock_min') is not None:
            query += " AND p.stock >= ?"
            params.append(filtros['stock_min'])

    c.execute(query + " ORDER BY p.nombre", params)
    datos = c.fetchall()
    conn.close()
    return datos

def actualizar_producto(producto_id, precio=None, stock=None):
    conn = conectar_bd()
    c = conn.cursor()
    campos = []
    valores = []
    if precio is not None:
        campos.append("precio = ?")
        valores.append(precio)
    if stock is not None:
        campos.append("stock = ?")
        valores.append(stock)
    if not campos:
        return False
    valores.append(producto_id)
    query = f"UPDATE productos SET {', '.join(campos)} WHERE id = ?"
    try:
        c.execute(query, valores)
        conn.commit()
        return True
    except sqlite3.Error:
        return False
    finally:
        conn.close()

def eliminar_producto(producto_id):
    conn = conectar_bd()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM productos WHERE id = ?", (producto_id,))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

# --- Ventas ---
def registrar_venta(cliente_id, carrito, descuento=0):
    """carrito es lista de (producto_id, cantidad, precio_unitario)"""
    conn = conectar_bd()
    c = conn.cursor()
    total_bruto = sum(cant * precio for _, cant, precio in carrito)
    total_neto = total_bruto - descuento
    if total_neto < 0:
        total_neto = 0

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        c.execute("INSERT INTO ventas (fecha, cliente_id, total, descuento) VALUES (?, ?, ?, ?)",
                  (fecha, cliente_id, total_neto, descuento))
        venta_id = c.lastrowid
        for prod_id, cant, precio in carrito:
            subtotal = cant * precio
            c.execute("INSERT INTO detalle_venta (venta_id, producto_id, cantidad, precio_unitario, subtotal) VALUES (?, ?, ?, ?, ?)",
                      (venta_id, prod_id, cant, precio, subtotal))
            # Reducir stock
            c.execute("UPDATE productos SET stock = stock - ? WHERE id = ? AND stock >= ?", (cant, prod_id, cant))
            if c.rowcount == 0:
                raise Exception("Stock insuficiente")
        conn.commit()
        return venta_id
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

def obtener_ventas(fecha_inicio=None, fecha_fin=None):
    conn = conectar_bd()
    c = conn.cursor()
    query = """
        SELECT v.id, v.fecha, c.nombre, v.total, v.descuento
        FROM ventas v
        LEFT JOIN clientes c ON v.cliente_id = c.id
        WHERE 1=1
    """
    params = []
    if fecha_inicio:
        query += " AND v.fecha >= ?"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND v.fecha <= ?"
        params.append(fecha_fin)
    c.execute(query + " ORDER BY v.fecha DESC", params)
    ventas = c.fetchall()
    conn.close()
    return ventas

def obtener_detalle_venta(venta_id):
    conn = conectar_bd()
    c = conn.cursor()
    c.execute("""
        SELECT p.nombre, dv.cantidad, dv.precio_unitario, dv.subtotal
        FROM detalle_venta dv
        JOIN productos p ON dv.producto_id = p.id
        WHERE dv.venta_id = ?
    """, (venta_id,))
    datos = c.fetchall()
    conn.close()
    return datos

# --- Compras a proveedores (reposición) ---
def registrar_compra(proveedor_id, items):
    """items es lista de (producto_id, cantidad, costo_unitario)"""
    conn = conectar_bd()
    c = conn.cursor()
    total = sum(cant * costo for _, cant, costo in items)
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        c.execute("INSERT INTO compras (fecha, proveedor_id, total) VALUES (?, ?, ?)",
                  (fecha, proveedor_id, total))
        compra_id = c.lastrowid
        for prod_id, cant, costo in items:
            subtotal = cant * costo
            c.execute("INSERT INTO detalle_compra (compra_id, producto_id, cantidad, costo_unitario, subtotal) VALUES (?, ?, ?, ?, ?)",
                      (compra_id, prod_id, cant, costo, subtotal))
            # Aumentar stock
            c.execute("UPDATE productos SET stock = stock + ? WHERE id = ?", (cant, prod_id))
        conn.commit()
        return compra_id
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

# --- Promociones/Descuentos ---
def agregar_promocion(nombre, tipo, valor, fecha_inicio, fecha_fin, producto_id=None):
    conn = conectar_bd()
    c = conn.cursor()
    try:
        c.execute("""INSERT INTO promociones (nombre, tipo, valor, fecha_inicio, fecha_fin, producto_id)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (nombre, tipo, valor, fecha_inicio, fecha_fin, producto_id))
        conn.commit()
        return True
    except sqlite3.Error:
        return False
    finally:
        conn.close()

def obtener_promociones_activas():
    """Devuelve promociones vigentes (fecha actual entre inicio y fin)"""
    hoy = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = conectar_bd()
    c = conn.cursor()
    c.execute("""
        SELECT id, nombre, tipo, valor, producto_id
        FROM promociones
        WHERE (fecha_inicio IS NULL OR fecha_inicio <= ?)
        AND (fecha_fin IS NULL OR fecha_fin >= ?)
    """, (hoy, hoy))
    datos = c.fetchall()
    conn.close()
    return datos

# ============================================
# FUNCIONES DE REPORTES (PDF, CSV, EXCEL)
# ============================================

def generar_factura_pdf(venta_id, ruta=None):
    """Genera un PDF con la factura de la venta"""
    if not ruta:
        ruta = f"factura_{venta_id}.pdf"
    ventas = obtener_ventas()
    venta = next((v for v in ventas if v[0] == venta_id), None)
    if not venta:
        return None
    detalles = obtener_detalle_venta(venta_id)

    c = canvas.Canvas(ruta, pagesize=letter)
    width, height = letter

    # Encabezado
    c.setFont("Helvetica-Bold", 16)
    c.drawString(1*inch, height-1*inch, "QuickMart Manager - FACTURA")
    c.setFont("Helvetica", 10)
    c.drawString(1*inch, height-1.5*inch, f"Venta #{venta_id}")
    c.drawString(1*inch, height-1.8*inch, f"Fecha: {venta[1]}")
    c.drawString(1*inch, height-2.1*inch, f"Cliente: {venta[2] if venta[2] else 'Genérico'}")
    c.drawString(1*inch, height-2.4*inch, f"Descuento: ${venta[4]:.2f}")

    # Tabla de productos
    data = [["Producto", "Cantidad", "Precio Unit.", "Subtotal"]]
    for det in detalles:
        data.append([det[0], str(det[1]), f"${det[2]:.2f}", f"${det[3]:.2f}"])
    data.append(["", "", "TOTAL:", f"${venta[3]:.2f}"])

    table = Table(data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-2), 1, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 1*inch, height-4.5*inch - (len(data)*0.2*inch))

    c.save()
    return ruta

def exportar_reporte_csv(data, columnas, nombre_archivo="reporte.csv"):
    """Exporta una lista de tuplas a CSV"""
    with open(nombre_archivo, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columnas)
        writer.writerows(data)
    return nombre_archivo

def exportar_reporte_excel(data, columnas, nombre_archivo="reporte.xlsx"):
    """Exporta una lista de tuplas a Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    # Encabezados
    for col, titulo in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal='center')
    # Datos
    for fila_idx, fila in enumerate(data, 2):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=fila_idx, column=col_idx, value=valor)
    wb.save(nombre_archivo)
    return nombre_archivo

# ============================================
# INTERFAZ GRÁFICA CON TKINTER
# ============================================

class LoginWindow:
    def __init__(self, master):
        self.master = master
        self.master.title("QuickMart - Login")
        self.master.geometry("300x150")
        self.master.resizable(False, False)

        self.frame = ttk.Frame(master, padding="10")
        self.frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        ttk.Label(self.frame, text="Usuario:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.entry_usuario = ttk.Entry(self.frame, width=25)
        self.entry_usuario.grid(row=0, column=1, pady=5)

        ttk.Label(self.frame, text="Contraseña:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.entry_password = ttk.Entry(self.frame, width=25, show="*")
        self.entry_password.grid(row=1, column=1, pady=5)

        self.btn_login = ttk.Button(self.frame, text="Iniciar Sesión", command=self.login)
        self.btn_login.grid(row=2, column=0, columnspan=2, pady=10)

        self.master.bind('<Return>', lambda e: self.login())

    def login(self):
        usuario = self.entry_usuario.get()
        password = self.entry_password.get()
        if not usuario or not password:
            messagebox.showerror("Error", "Ingrese usuario y contraseña")
            return
        user_id, rol = autenticar_usuario(usuario, password)
        if user_id:
            self.master.destroy()
            root = tk.Tk()
            app = QuickMartApp(root, user_id, rol)
            root.mainloop()
        else:
            messagebox.showerror("Error", "Credenciales incorrectas")

class QuickMartApp:
    def __init__(self, master, user_id, rol):
        self.master = master
        self.user_id = user_id
        self.rol = rol
        self.master.title(f"QuickMart Manager PRO - {rol.capitalize()}")
        self.master.geometry("900x600")

        # Crear menú
        menubar = tk.Menu(self.master)
        self.master.config(menu=menubar)

        # Menú Archivo
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=file_menu)
        file_menu.add_command(label="Cambiar contraseña", command=self.cambiar_password)
        file_menu.add_separator()
        file_menu.add_command(label="Salir", command=self.master.quit)

        # Menú Reportes
        report_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Reportes", menu=report_menu)
        report_menu.add_command(label="Exportar productos a CSV", command=self.exportar_productos_csv)
        report_menu.add_command(label="Exportar productos a Excel", command=self.exportar_productos_excel)
        report_menu.add_command(label="Exportar ventas a CSV", command=self.exportar_ventas_csv)
        report_menu.add_command(label="Exportar ventas a Excel", command=self.exportar_ventas_excel)

        # Panel de navegación (notebook)
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Pestañas
        self.tab_productos = ttk.Frame(self.notebook)
        self.tab_proveedores = ttk.Frame(self.notebook)
        self.tab_clientes = ttk.Frame(self.notebook)
        self.tab_ventas = ttk.Frame(self.notebook)
        self.tab_compras = ttk.Frame(self.notebook)
        self.tab_promociones = ttk.Frame(self.notebook)
        self.tab_reportes = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_productos, text="Productos")
        self.notebook.add(self.tab_proveedores, text="Proveedores")
        self.notebook.add(self.tab_clientes, text="Clientes")
        self.notebook.add(self.tab_ventas, text="Ventas")
        self.notebook.add(self.tab_compras, text="Compras")
        self.notebook.add(self.tab_promociones, text="Promociones")
        self.notebook.add(self.tab_reportes, text="Reportes")

        # Inicializar cada pestaña
        self.init_tab_productos()
        self.init_tab_proveedores()
        self.init_tab_clientes()
        self.init_tab_ventas()
        self.init_tab_compras()
        self.init_tab_promociones()
        self.init_tab_reportes()

        # Cargar datos iniciales
        self.cargar_productos()
        self.cargar_proveedores()
        self.cargar_clientes()
        self.cargar_ventas()
        self.cargar_compras()
        self.cargar_promociones()

    # ---------- Pestaña Productos ----------
    def init_tab_productos(self):
        frame = self.tab_productos
        # Filtros
        filtros_frame = ttk.LabelFrame(frame, text="Filtros de búsqueda avanzada", padding=5)
        filtros_frame.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=5, padx=5)

        ttk.Label(filtros_frame, text="Nombre:").grid(row=0, column=0, padx=5)
        self.filtro_nombre = ttk.Entry(filtros_frame, width=15)
        self.filtro_nombre.grid(row=0, column=1, padx=5)

        ttk.Label(filtros_frame, text="Categoría:").grid(row=0, column=2, padx=5)
        self.filtro_categoria = ttk.Entry(filtros_frame, width=15)
        self.filtro_categoria.grid(row=0, column=3, padx=5)

        ttk.Label(filtros_frame, text="Proveedor ID:").grid(row=0, column=4, padx=5)
        self.filtro_proveedor = ttk.Entry(filtros_frame, width=10)
        self.filtro_proveedor.grid(row=0, column=5, padx=5)

        ttk.Label(filtros_frame, text="Precio min:").grid(row=1, column=0, padx=5)
        self.filtro_precio_min = ttk.Entry(filtros_frame, width=10)
        self.filtro_precio_min.grid(row=1, column=1, padx=5)

        ttk.Label(filtros_frame, text="Precio max:").grid(row=1, column=2, padx=5)
        self.filtro_precio_max = ttk.Entry(filtros_frame, width=10)
        self.filtro_precio_max.grid(row=1, column=3, padx=5)

        ttk.Label(filtros_frame, text="Stock min:").grid(row=1, column=4, padx=5)
        self.filtro_stock_min = ttk.Entry(filtros_frame, width=10)
        self.filtro_stock_min.grid(row=1, column=5, padx=5)

        ttk.Button(filtros_frame, text="Filtrar", command=self.cargar_productos).grid(row=2, column=0, columnspan=2, pady=5)
        ttk.Button(filtros_frame, text="Limpiar filtros", command=self.limpiar_filtros_productos).grid(row=2, column=2, columnspan=2, pady=5)

        # Tabla de productos
        columns = ("ID", "Código", "Nombre", "Precio", "Stock", "Categoría", "Proveedor")
        self.tree_productos = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            self.tree_productos.heading(col, text=col)
            if col in ("ID", "Precio", "Stock"):
                self.tree_productos.column(col, width=70)
            elif col == "Código":
                self.tree_productos.column(col, width=80)
            elif col == "Nombre":
                self.tree_productos.column(col, width=150)
            else:
                self.tree_productos.column(col, width=100)
        self.tree_productos.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree_productos.yview)
        scrollbar.grid(row=1, column=4, sticky=(tk.N, tk.S))
        self.tree_productos.configure(yscrollcommand=scrollbar.set)

        # Botones de acciones
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=4, pady=5)

        ttk.Button(btn_frame, text="Agregar", command=self.agregar_producto).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Editar (precio/stock)", command=self.editar_producto).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Eliminar", command=self.eliminar_producto).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Generar factura (seleccionar venta)", command=self.generar_factura_desde_producto).pack(side=tk.LEFT, padx=5)

        frame.rowconfigure(1, weight=1)
        frame.columnconfigure(0, weight=1)

    def limpiar_filtros_productos(self):
        self.filtro_nombre.delete(0, tk.END)
        self.filtro_categoria.delete(0, tk.END)
        self.filtro_proveedor.delete(0, tk.END)
        self.filtro_precio_min.delete(0, tk.END)
        self.filtro_precio_max.delete(0, tk.END)
        self.filtro_stock_min.delete(0, tk.END)
        self.cargar_productos()

    def cargar_productos(self):
        for item in self.tree_productos.get_children():
            self.tree_productos.delete(item)
        filtros = {}
        if self.filtro_nombre.get():
            filtros['nombre'] = self.filtro_nombre.get()
        if self.filtro_categoria.get():
            filtros['categoria'] = self.filtro_categoria.get()
        if self.filtro_proveedor.get().isdigit():
            filtros['proveedor_id'] = int(self.filtro_proveedor.get())
        if self.filtro_precio_min.get().replace('.', '', 1).isdigit():
            filtros['precio_min'] = float(self.filtro_precio_min.get())
        if self.filtro_precio_max.get().replace('.', '', 1).isdigit():
            filtros['precio_max'] = float(self.filtro_precio_max.get())
        if self.filtro_stock_min.get().isdigit():
            filtros['stock_min'] = int(self.filtro_stock_min.get())

        productos = obtener_productos(filtros)
        for p in productos:
            self.tree_productos.insert("", tk.END, values=p)

    def agregar_producto(self):
        # Ventana emergente para agregar producto
        dialog = tk.Toplevel(self.master)
        dialog.title("Agregar Producto")
        dialog.geometry("300x400")
        dialog.transient(self.master)
        dialog.grab_set()

        ttk.Label(dialog, text="Código:").grid(row=0, column=0, padx=5, pady=5)
        entry_codigo = ttk.Entry(dialog)
        entry_codigo.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Nombre:").grid(row=1, column=0, padx=5, pady=5)
        entry_nombre = ttk.Entry(dialog)
        entry_nombre.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Descripción:").grid(row=2, column=0, padx=5, pady=5)
        entry_desc = ttk.Entry(dialog)
        entry_desc.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Precio:").grid(row=3, column=0, padx=5, pady=5)
        entry_precio = ttk.Entry(dialog)
        entry_precio.grid(row=3, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Stock:").grid(row=4, column=0, padx=5, pady=5)
        entry_stock = ttk.Entry(dialog)
        entry_stock.grid(row=4, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Categoría:").grid(row=5, column=0, padx=5, pady=5)
        entry_cat = ttk.Entry(dialog)
        entry_cat.grid(row=5, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Proveedor ID:").grid(row=6, column=0, padx=5, pady=5)
        entry_prov = ttk.Entry(dialog)
        entry_prov.grid(row=6, column=1, padx=5, pady=5)

        def guardar():
            codigo = entry_codigo.get().strip()
            nombre = entry_nombre.get().strip()
            if not codigo or not nombre:
                messagebox.showerror("Error", "Código y nombre son obligatorios")
                return
            try:
                precio = float(entry_precio.get())
                stock = int(entry_stock.get())
            except ValueError:
                messagebox.showerror("Error", "Precio y stock deben ser números")
                return
            desc = entry_desc.get().strip()
            cat = entry_cat.get().strip()
            prov_id = entry_prov.get().strip()
            if prov_id.isdigit():
                prov_id = int(prov_id)
            else:
                prov_id = None
            if agregar_producto(codigo, nombre, desc, precio, stock, cat, prov_id):
                messagebox.showinfo("Éxito", "Producto agregado")
                dialog.destroy()
                self.cargar_productos()
            else:
                messagebox.showerror("Error", "Código duplicado o error en BD")

        ttk.Button(dialog, text="Guardar", command=guardar).grid(row=7, column=0, columnspan=2, pady=10)

    def editar_producto(self):
        seleccion = self.tree_productos.selection()
        if not seleccion:
            messagebox.showwarning("Seleccionar", "Selecciona un producto")
            return
        item = self.tree_productos.item(seleccion[0])
        prod_id = item['values'][0]
        precio_actual = item['values'][3]
        stock_actual = item['values'][4]

        nuevo_precio = simpledialog.askfloat("Editar precio", "Nuevo precio:", initialvalue=precio_actual)
        nuevo_stock = simpledialog.askinteger("Editar stock", "Nuevo stock:", initialvalue=stock_actual)
        if nuevo_precio is not None and nuevo_stock is not None:
            if actualizar_producto(prod_id, precio=nuevo_precio, stock=nuevo_stock):
                messagebox.showinfo("Éxito", "Producto actualizado")
                self.cargar_productos()
            else:
                messagebox.showerror("Error", "No se pudo actualizar")

    def eliminar_producto(self):
        seleccion = self.tree_productos.selection()
        if not seleccion:
            messagebox.showwarning("Seleccionar", "Selecciona un producto")
            return
        item = self.tree_productos.item(seleccion[0])
        prod_id = item['values'][0]
        if messagebox.askyesno("Confirmar", "¿Eliminar este producto?"):
            if eliminar_producto(prod_id):
                messagebox.showinfo("Éxito", "Producto eliminado")
                self.cargar_productos()
            else:
                messagebox.showerror("Error", "No se puede eliminar porque tiene ventas asociadas")

    def generar_factura_desde_producto(self):
        # Esto es un placeholder; la factura se genera desde la pestaña de ventas
        messagebox.showinfo("Info", "Ve a la pestaña Ventas, selecciona una venta y usa 'Generar Factura'")

    # ---------- Pestaña Proveedores ----------
    def init_tab_proveedores(self):
        frame = self.tab_proveedores
        columns = ("ID", "Nombre", "Teléfono", "Email", "Dirección")
        self.tree_proveedores = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            self.tree_proveedores.heading(col, text=col)
            self.tree_proveedores.column(col, width=100 if col != "Dirección" else 150)
        self.tree_proveedores.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="Agregar", command=self.agregar_proveedor).pack(side=tk.LEFT, padx=5)

    def cargar_proveedores(self):
        for item in self.tree_proveedores.get_children():
            self.tree_proveedores.delete(item)
        for p in obtener_proveedores():
            self.tree_proveedores.insert("", tk.END, values=p)

    def agregar_proveedor(self):
        dialog = tk.Toplevel(self.master)
        dialog.title("Agregar Proveedor")
        dialog.geometry("300x250")
        dialog.transient(self.master)
        dialog.grab_set()

        ttk.Label(dialog, text="Nombre:").grid(row=0, column=0, padx=5, pady=5)
        entry_nombre = ttk.Entry(dialog)
        entry_nombre.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Teléfono:").grid(row=1, column=0, padx=5, pady=5)
        entry_tel = ttk.Entry(dialog)
        entry_tel.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Email:").grid(row=2, column=0, padx=5, pady=5)
        entry_email = ttk.Entry(dialog)
        entry_email.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Dirección:").grid(row=3, column=0, padx=5, pady=5)
        entry_dir = ttk.Entry(dialog)
        entry_dir.grid(row=3, column=1, padx=5, pady=5)

        def guardar():
            nombre = entry_nombre.get().strip()
            if not nombre:
                messagebox.showerror("Error", "Nombre obligatorio")
                return
            if agregar_proveedor(nombre, entry_tel.get(), entry_email.get(), entry_dir.get()):
                messagebox.showinfo("Éxito", "Proveedor agregado")
                dialog.destroy()
                self.cargar_proveedores()
            else:
                messagebox.showerror("Error", "Error al agregar")
        ttk.Button(dialog, text="Guardar", command=guardar).grid(row=4, column=0, columnspan=2, pady=10)

    # ---------- Pestaña Clientes ----------
    def init_tab_clientes(self):
        frame = self.tab_clientes
        columns = ("ID", "Nombre", "Teléfono", "Email", "Dirección")
        self.tree_clientes = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            self.tree_clientes.heading(col, text=col)
            self.tree_clientes.column(col, width=100 if col != "Dirección" else 150)
        self.tree_clientes.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="Agregar", command=self.agregar_cliente).pack(side=tk.LEFT, padx=5)

    def cargar_clientes(self):
        for item in self.tree_clientes.get_children():
            self.tree_clientes.delete(item)
        for c in obtener_clientes():
            self.tree_clientes.insert("", tk.END, values=c)

    def agregar_cliente(self):
        dialog = tk.Toplevel(self.master)
        dialog.title("Agregar Cliente")
        dialog.geometry("300x250")
        dialog.transient(self.master)
        dialog.grab_set()

        ttk.Label(dialog, text="Nombre:").grid(row=0, column=0, padx=5, pady=5)
        entry_nombre = ttk.Entry(dialog)
        entry_nombre.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Teléfono:").grid(row=1, column=0, padx=5, pady=5)
        entry_tel = ttk.Entry(dialog)
        entry_tel.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Email:").grid(row=2, column=0, padx=5, pady=5)
        entry_email = ttk.Entry(dialog)
        entry_email.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Dirección:").grid(row=3, column=0, padx=5, pady=5)
        entry_dir = ttk.Entry(dialog)
        entry_dir.grid(row=3, column=1, padx=5, pady=5)

        def guardar():
            nombre = entry_nombre.get().strip()
            if not nombre:
                messagebox.showerror("Error", "Nombre obligatorio")
                return
            if agregar_cliente(nombre, entry_tel.get(), entry_email.get(), entry_dir.get()):
                messagebox.showinfo("Éxito", "Cliente agregado")
                dialog.destroy()
                self.cargar_clientes()
            else:
                messagebox.showerror("Error", "Error al agregar")
        ttk.Button(dialog, text="Guardar", command=guardar).grid(row=4, column=0, columnspan=2, pady=10)

    # ---------- Pestaña Ventas ----------
    def init_tab_ventas(self):
        frame = self.tab_ventas
        # Frame superior para controles
        top_frame = ttk.Frame(frame)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(top_frame, text="Cliente ID:").pack(side=tk.LEFT, padx=5)
        self.entry_cliente_venta = ttk.Entry(top_frame, width=10)
        self.entry_cliente_venta.pack(side=tk.LEFT, padx=5)

        ttk.Label(top_frame, text="Descuento ($):").pack(side=tk.LEFT, padx=5)
        self.entry_descuento = ttk.Entry(top_frame, width=10)
        self.entry_descuento.pack(side=tk.LEFT, padx=5)

        ttk.Button(top_frame, text="Nueva Venta", command=self.nueva_venta).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Generar Factura", command=self.generar_factura).pack(side=tk.LEFT, padx=5)

        # Tabla de ventas
        columns = ("ID", "Fecha", "Cliente", "Total", "Descuento")
        self.tree_ventas = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            self.tree_ventas.heading(col, text=col)
            self.tree_ventas.column(col, width=100 if col != "Cliente" else 150)
        self.tree_ventas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def cargar_ventas(self):
        for item in self.tree_ventas.get_children():
            self.tree_ventas.delete(item)
        for v in obtener_ventas():
            self.tree_ventas.insert("", tk.END, values=v)

    def nueva_venta(self):
        # Ventana para agregar productos al carrito
        dialog = tk.Toplevel(self.master)
        dialog.title("Nueva Venta")
        dialog.geometry("500x400")
        dialog.transient(self.master)
        dialog.grab_set()

        # Lista de productos disponibles
        ttk.Label(dialog, text="Productos disponibles (doble click para agregar)").pack(pady=5)
        tree = ttk.Treeview(dialog, columns=("ID", "Código", "Nombre", "Precio", "Stock"), show="headings")
        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        for col in tree['columns']:
            tree.heading(col, text=col)
            tree.column(col, width=80)

        # Cargar productos
        for p in obtener_productos():
            tree.insert("", tk.END, values=p[:5])

        # Carrito (se almacenará en una lista)
        carrito = []

        def agregar_al_carrito(event):
            seleccion = tree.selection()
            if not seleccion:
                return
            item = tree.item(seleccion[0])
            prod_id = item['values'][0]
            nombre = item['values'][2]
            precio = item['values'][3]
            stock = item['values'][4]
            cantidad = simpledialog.askinteger("Cantidad", f"Cantidad de {nombre}:", minvalue=1, maxvalue=stock)
            if cantidad:
                carrito.append((prod_id, cantidad, precio))
                messagebox.showinfo("Agregado", f"{nombre} x {cantidad} agregado al carrito")

        tree.bind("<Double-Button-1>", agregar_al_carrito)

        # Mostrar carrito
        ttk.Label(dialog, text="Carrito actual:").pack(pady=5)
        carrito_listbox = tk.Listbox(dialog, height=5)
        carrito_listbox.pack(fill=tk.X, padx=5)

        def actualizar_carrito():
            carrito_listbox.delete(0, tk.END)
            for prod_id, cant, precio in carrito:
                nombre = next((p[2] for p in obtener_productos() if p[0] == prod_id), "Desconocido")
                carrito_listbox.insert(tk.END, f"{nombre} x {cant} = ${cant*precio:.2f}")

        def finalizar_venta():
            if not carrito:
                messagebox.showwarning("Carrito vacío", "Agregue productos")
                return
            cliente_id = self.entry_cliente_venta.get()
            if cliente_id.isdigit():
                cliente_id = int(cliente_id)
            else:
                cliente_id = None
            descuento = self.entry_descuento.get()
            try:
                descuento = float(descuento) if descuento else 0.0
            except ValueError:
                descuento = 0.0
            try:
                venta_id = registrar_venta(cliente_id, carrito, descuento)
                messagebox.showinfo("Éxito", f"Venta #{venta_id} registrada")
                dialog.destroy()
                self.cargar_ventas()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo registrar: {e}")

        ttk.Button(dialog, text="Finalizar Venta", command=finalizar_venta).pack(pady=10)

    def generar_factura(self):
        seleccion = self.tree_ventas.selection()
        if not seleccion:
            messagebox.showwarning("Seleccionar", "Selecciona una venta")
            return
        item = self.tree_ventas.item(seleccion[0])
        venta_id = item['values'][0]
        ruta = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if ruta:
            try:
                archivo = generar_factura_pdf(venta_id, ruta)
                messagebox.showinfo("Factura", f"Factura generada: {archivo}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo generar PDF: {e}")

    # ---------- Pestaña Compras (reposición) ----------
    def init_tab_compras(self):
        frame = self.tab_compras
        top_frame = ttk.Frame(frame)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(top_frame, text="Proveedor ID:").pack(side=tk.LEFT, padx=5)
        self.entry_proveedor_compra = ttk.Entry(top_frame, width=10)
        self.entry_proveedor_compra.pack(side=tk.LEFT, padx=5)

        ttk.Button(top_frame, text="Nueva Compra", command=self.nueva_compra).pack(side=tk.LEFT, padx=5)

        columns = ("ID", "Fecha", "Proveedor", "Total")
        self.tree_compras = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            self.tree_compras.heading(col, text=col)
            self.tree_compras.column(col, width=100)
        self.tree_compras.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def cargar_compras(self):
        for item in self.tree_compras.get_children():
            self.tree_compras.delete(item)
        conn = conectar_bd()
        c = conn.cursor()
        c.execute("SELECT c.id, c.fecha, p.nombre, c.total FROM compras c LEFT JOIN proveedores p ON c.proveedor_id = p.id ORDER BY c.fecha DESC")
        datos = c.fetchall()
        conn.close()
        for row in datos:
            self.tree_compras.insert("", tk.END, values=row)

    def nueva_compra(self):
        dialog = tk.Toplevel(self.master)
        dialog.title("Nueva Compra a Proveedor")
        dialog.geometry("500x400")
        dialog.transient(self.master)
        dialog.grab_set()

        # Lista de productos
        ttk.Label(dialog, text="Selecciona producto (doble click)").pack(pady=5)
        tree = ttk.Treeview(dialog, columns=("ID", "Código", "Nombre", "Stock actual"), show="headings")
        tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        for col in tree['columns']:
            tree.heading(col, text=col)
            tree.column(col, width=80)
        for p in obtener_productos():
            tree.insert("", tk.END, values=(p[0], p[1], p[2], p[5]))

        items = []  # (producto_id, cantidad, costo_unitario)

        def agregar_item(event):
            seleccion = tree.selection()
            if not seleccion:
                return
            item = tree.item(seleccion[0])
            prod_id = item['values'][0]
            nombre = item['values'][2]
            cantidad = simpledialog.askinteger("Cantidad", f"Cantidad de {nombre} a comprar:", minvalue=1)
            if cantidad:
                costo = simpledialog.askfloat("Costo unitario", f"Costo unitario de {nombre}:", minvalue=0)
                if costo is not None:
                    items.append((prod_id, cantidad, costo))
                    messagebox.showinfo("Agregado", f"{nombre} x {cantidad} a costo {costo:.2f}")

        tree.bind("<Double-Button-1>", agregar_item)

        ttk.Label(dialog, text="Items agregados:").pack(pady=5)
        listbox = tk.Listbox(dialog, height=5)
        listbox.pack(fill=tk.X, padx=5)

        def actualizar_lista():
            listbox.delete(0, tk.END)
            for prod_id, cant, costo in items:
                nombre = next((p[2] for p in obtener_productos() if p[0] == prod_id), "Desconocido")
                listbox.insert(tk.END, f"{nombre} x {cant} = ${cant*costo:.2f}")

        def finalizar_compra():
            if not items:
                messagebox.showwarning("Sin items", "Agregue productos")
                return
            proveedor_id = self.entry_proveedor_compra.get()
            if proveedor_id.isdigit():
                proveedor_id = int(proveedor_id)
            else:
                proveedor_id = None
            try:
                compra_id = registrar_compra(proveedor_id, items)
                messagebox.showinfo("Éxito", f"Compra #{compra_id} registrada")
                dialog.destroy()
                self.cargar_compras()
                self.cargar_productos()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo registrar: {e}")

        ttk.Button(dialog, text="Finalizar Compra", command=finalizar_compra).pack(pady=10)

    # ---------- Pestaña Promociones ----------
    def init_tab_promociones(self):
        frame = self.tab_promociones
        top_frame = ttk.Frame(frame)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(top_frame, text="Agregar Promoción", command=self.agregar_promocion).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Ver promociones activas", command=self.ver_promociones_activas).pack(side=tk.LEFT, padx=5)

        columns = ("ID", "Nombre", "Tipo", "Valor", "Producto ID")
        self.tree_promociones = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            self.tree_promociones.heading(col, text=col)
            self.tree_promociones.column(col, width=100)
        self.tree_promociones.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def cargar_promociones(self):
        for item in self.tree_promociones.get_children():
            self.tree_promociones.delete(item)
        conn = conectar_bd()
        c = conn.cursor()
        c.execute("SELECT id, nombre, tipo, valor, producto_id FROM promociones")
        for row in c.fetchall():
            self.tree_promociones.insert("", tk.END, values=row)
        conn.close()

    def agregar_promocion(self):
        dialog = tk.Toplevel(self.master)
        dialog.title("Agregar Promoción")
        dialog.geometry("300x300")
        dialog.transient(self.master)
        dialog.grab_set()

        ttk.Label(dialog, text="Nombre:").grid(row=0, column=0, padx=5, pady=5)
        entry_nombre = ttk.Entry(dialog)
        entry_nombre.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Tipo (porcentaje/fijo):").grid(row=1, column=0, padx=5, pady=5)
        entry_tipo = ttk.Combobox(dialog, values=["porcentaje", "fijo"])
        entry_tipo.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Valor:").grid(row=2, column=0, padx=5, pady=5)
        entry_valor = ttk.Entry(dialog)
        entry_valor.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Fecha inicio (YYYY-MM-DD HH:MM:SS):").grid(row=3, column=0, padx=5, pady=5)
        entry_inicio = ttk.Entry(dialog)
        entry_inicio.grid(row=3, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Fecha fin:").grid(row=4, column=0, padx=5, pady=5)
        entry_fin = ttk.Entry(dialog)
        entry_fin.grid(row=4, column=1, padx=5, pady=5)

        ttk.Label(dialog, text="Producto ID (opcional):").grid(row=5, column=0, padx=5, pady=5)
        entry_prod = ttk.Entry(dialog)
        entry_prod.grid(row=5, column=1, padx=5, pady=5)

        def guardar():
            nombre = entry_nombre.get().strip()
            tipo = entry_tipo.get().strip()
            if not nombre or not tipo:
                messagebox.showerror("Error", "Nombre y tipo obligatorios")
                return
            try:
                valor = float(entry_valor.get())
            except ValueError:
                messagebox.showerror("Error", "Valor debe ser número")
                return
            inicio = entry_inicio.get().strip() or None
            fin = entry_fin.get().strip() or None
            prod_id = entry_prod.get().strip()
            if prod_id.isdigit():
                prod_id = int(prod_id)
            else:
                prod_id = None
            if agregar_promocion(nombre, tipo, valor, inicio, fin, prod_id):
                messagebox.showinfo("Éxito", "Promoción agregada")
                dialog.destroy()
                self.cargar_promociones()
            else:
                messagebox.showerror("Error", "No se pudo agregar")
        ttk.Button(dialog, text="Guardar", command=guardar).grid(row=6, column=0, columnspan=2, pady=10)

    def ver_promociones_activas(self):
        activas = obtener_promociones_activas()
        if not activas:
            messagebox.showinfo("Promociones", "No hay promociones activas")
        else:
            mensaje = "Promociones activas:\n"
            for p in activas:
                mensaje += f"{p[1]} - {p[2]} {p[3]} (Producto ID: {p[4] if p[4] else 'Todos'})\n"
            messagebox.showinfo("Promociones activas", mensaje)

    # ---------- Pestaña Reportes ----------
    def init_tab_reportes(self):
        frame = self.tab_reportes
        ttk.Label(frame, text="Reportes y exportaciones", font=("Helvetica", 14)).pack(pady=10)

        ttk.Button(frame, text="Exportar productos a CSV", command=self.exportar_productos_csv).pack(pady=5)
        ttk.Button(frame, text="Exportar productos a Excel", command=self.exportar_productos_excel).pack(pady=5)
        ttk.Button(frame, text="Exportar ventas a CSV", command=self.exportar_ventas_csv).pack(pady=5)
        ttk.Button(frame, text="Exportar ventas a Excel", command=self.exportar_ventas_excel).pack(pady=5)
        ttk.Button(frame, text="Ver ventas del día (popup)", command=self.ventas_dia).pack(pady=5)

    def exportar_productos_csv(self):
        productos = obtener_productos()
        if not productos:
            messagebox.showinfo("Sin datos", "No hay productos para exportar")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if ruta:
            columnas = ["ID", "Código", "Nombre", "Descripción", "Precio", "Stock", "Categoría", "Proveedor"]
            exportar_reporte_csv(productos, columnas, ruta)
            messagebox.showinfo("Exportado", f"Exportado a {ruta}")

    def exportar_productos_excel(self):
        productos = obtener_productos()
        if not productos:
            messagebox.showinfo("Sin datos", "No hay productos para exportar")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if ruta:
            columnas = ["ID", "Código", "Nombre", "Descripción", "Precio", "Stock", "Categoría", "Proveedor"]
            exportar_reporte_excel(productos, columnas, ruta)
            messagebox.showinfo("Exportado", f"Exportado a {ruta}")

    def exportar_ventas_csv(self):
        ventas = obtener_ventas()
        if not ventas:
            messagebox.showinfo("Sin datos", "No hay ventas para exportar")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if ruta:
            columnas = ["ID", "Fecha", "Cliente", "Total", "Descuento"]
            exportar_reporte_csv(ventas, columnas, ruta)
            messagebox.showinfo("Exportado", f"Exportado a {ruta}")

    def exportar_ventas_excel(self):
        ventas = obtener_ventas()
        if not ventas:
            messagebox.showinfo("Sin datos", "No hay ventas para exportar")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if ruta:
            columnas = ["ID", "Fecha", "Cliente", "Total", "Descuento"]
            exportar_reporte_excel(ventas, columnas, ruta)
            messagebox.showinfo("Exportado", f"Exportado a {ruta}")

    def ventas_dia(self):
        hoy = datetime.now().strftime("%Y-%m-%d")
        ventas = obtener_ventas(fecha_inicio=hoy, fecha_fin=hoy + " 23:59:59")
        total = sum(v[3] for v in ventas)
        messagebox.showinfo("Ventas del día", f"Ventas del {hoy}:\nTotal: ${total:.2f}\nCantidad: {len(ventas)}")

    # ---------- Funciones de usuario ----------
    def cambiar_password(self):
        nueva = simpledialog.askstring("Cambiar contraseña", "Nueva contraseña:", show="*")
        if nueva:
            cambiar_contrasena(self.user_id, nueva)
            messagebox.showinfo("Éxito", "Contraseña actualizada")

# ============================================
# PUNTO DE ENTRADA
# ============================================

if __name__ == "__main__":
    # Crear tablas antes de iniciar GUI
    crear_tablas()
    # Iniciar ventana de login
    root = tk.Tk()
    login = LoginWindow(root)
    root.mainloop()
